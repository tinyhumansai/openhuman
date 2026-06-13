"""
工程款利息核算台账生成器
接受 JSON 数据，生成 .xlsx 格式的专业台账

用法:
  python generate_ledger.py --input payment_data.json --output 工程款台账.xlsx

输入 JSON 格式:
{
  "project_name": "XX项目",
  "contract_no": "2024-XX-001",
  "periods": [
    {
      "period": 1,
      "period_range": "2024.01-2024.02",
      "approved_amount": 5000000,
      "deductions": {
        "retention": 150000,
        "owner_supplied": 0,
        "utilities": 5000,
        "penalties": 0,
        "other": 0
      },
      "net_payable": 4845000,
      "due_date": "2024-03-20",
      "conditions_met_date": "2024-03-18",
      "actual_paid_date": "2024-08-15",
      "actual_paid_amount": 4845000,
      "interest_segments": [
        {"start": "2024-03-21", "end": "2024-06-19", "days": 91, "rate": 3.45, "interest": 41625.21},
        {"start": "2024-06-20", "end": "2024-07-21", "days": 32, "rate": 3.45, "interest": 14637.53},
        {"start": "2024-07-22", "end": "2024-08-14", "days": 24, "rate": 3.35, "interest": 10684.93}
      ],
      "total_interest": 66947.67
    }
  ]
}
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from utils import format_cny

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl 包: pip3 install openpyxl")
    sys.exit(1)


# ── 样式定义 ────────────────────────────────────────────────────

HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="PingFang SC", size=10)
CNY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00"%"'
DATE_FORMAT = 'YYYY-MM-DD'
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
HIGH_RISK_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def create_ledger(data: dict, output_path: str):
    """生成工程款核算台账 Excel"""

    wb = openpyxl.Workbook()

    # ── Sheet 1: 核算台账 ──
    ws = wb.active
    ws.title = "工程款核算台账"

    # 标题行
    project_name = data.get("project_name", "工程项目")
    ws.merge_cells("A1:R1")
    ws["A1"] = f"{project_name} — 进度款及逾期利息核算台账"
    ws["A1"].font = Font(name="PingFang SC", size=14, bold=True, color="2F5496")
    ws["A1"].alignment = CENTER_ALIGN

    # 合同信息行
    ws.merge_cells("A2:R2")
    ws["A2"] = f"合同编号: {data.get('contract_no', '')}  |  制表日期: {date.today().isoformat()}"
    ws["A2"].font = Font(name="PingFang SC", size=9, color="808080")
    ws["A2"].alignment = CENTER_ALIGN

    # 表头（第 4 行）
    headers = [
        "期次", "施工期间", "获批金额", "质保金暂扣", "甲供材扣款",
        "水电费", "罚款/其他", "应付净额", "约定付款日",
        "前置条件完成日", "实际付款日", "延迟天数",
        "计息起始日", "计息终止日", "LPR(%)", "分段天数",
        "当期利息", "已付本金", "累计欠付", "备注"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据行
    row = 5
    total_interest = 0.0
    total_approved = 0.0
    total_net = 0.0

    for period in data.get("periods", []):
        p = period
        deductions = p.get("deductions", {})
        net = p.get("net_payable", 0)
        total_approved += p.get("approved_amount", 0)
        total_net += net
        period_interest = p.get("total_interest", 0)
        total_interest += period_interest

        values = [
            f"第{p['period']}期",
            p.get("period_range", ""),
            p.get("approved_amount", 0),
            deductions.get("retention", 0),
            deductions.get("owner_supplied", 0),
            deductions.get("utilities", 0),
            deductions.get("penalties", 0) + deductions.get("other", 0),
            net,
            p.get("due_date", ""),
            p.get("conditions_met_date", ""),
            p.get("actual_paid_date", ""),
            p.get("delay_days", ""),
            "",  # 计息起始日 (从 interest_segments 填充)
            "",  # 计息终止日
            "",  # LPR
            "",  # 分段天数
            period_interest,
            p.get("actual_paid_amount", 0),
            "",  # 累计欠付
            p.get("notes", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col in (3, 4, 5, 6, 7, 8, 17, 18):
                cell.alignment = RIGHT_ALIGN
                if isinstance(val, (int, float)):
                    cell.number_format = CNY_FORMAT
            elif col in (12, 16):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

        row += 1

    # ── 汇总行 ──
    summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    summary_font = Font(name="PingFang SC", size=10, bold=True)

    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value="合计").font = summary_font
    ws.cell(row=row, column=3, value=total_approved).number_format = CNY_FORMAT
    ws.cell(row=row, column=8, value=total_net).number_format = CNY_FORMAT
    ws.cell(row=row, column=17, value=total_interest).number_format = CNY_FORMAT
    for col in range(1, 21):
        ws.cell(row=row, column=col).fill = summary_fill
        ws.cell(row=row, column=col).font = summary_font
        ws.cell(row=row, column=col).border = THIN_BORDER

    # ── 列宽 ──
    col_widths = [8, 16, 14, 12, 12, 10, 10, 14, 14, 14, 14, 10, 14, 14, 10, 10, 14, 14, 14, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结表头
    ws.freeze_panes = "A5"

    # ── Sheet 2: 分段计息明细 ──
    ws2 = wb.create_sheet("分段计息明细")
    ws2.append(["期次", "分段起始日", "分段终止日", "天数", "LPR(%)", "计息本金", "应计利息", "累计利息"])

    detail_row = 2
    running_interest = 0.0
    for period in data.get("periods", []):
        net = period.get("net_payable", 0)
        for seg in period.get("interest_segments", []):
            seg_interest = seg.get("interest", 0)
            running_interest += seg_interest
            ws2.append([
                f"第{period['period']}期",
                seg.get("start", ""),
                seg.get("end", ""),
                seg.get("days", 0),
                seg.get("rate", 0),
                net,
                seg_interest,
                running_interest,
            ])
            # 样式
            for col in range(1, 9):
                ws2.cell(row=detail_row, column=col).font = DATA_FONT
                ws2.cell(row=detail_row, column=col).border = THIN_BORDER
                if col in (6, 7, 8):
                    ws2.cell(row=detail_row, column=col).number_format = CNY_FORMAT
            detail_row += 1

    # Sheet 2 样式
    for col in range(1, 9):
        ws2.cell(row=1, column=col).font = HEADER_FONT
        ws2.cell(row=1, column=col).fill = HEADER_FILL
        ws2.cell(row=1, column=col).border = THIN_BORDER
    ws2.freeze_panes = "A2"
    for col, w in enumerate([8, 14, 14, 8, 10, 14, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: 摘要信息 ──
    ws3 = wb.create_sheet("摘要")
    info = [
        ("项目名称", data.get("project_name", "")),
        ("合同编号", data.get("contract_no", "")),
        ("核算期数", f"{len(data.get('periods', []))} 期"),
        ("合计获批金额", format_cny(total_approved)),
        ("合计应付净额", format_cny(total_net)),
        ("合计应付利息", format_cny(total_interest)),
        ("制表日期", date.today().isoformat()),
    ]
    for i, (label, value) in enumerate(info, 1):
        ws3.cell(row=i, column=1, value=label).font = Font(name="PingFang SC", size=10, bold=True)
        ws3.cell(row=i, column=2, value=value).font = DATA_FONT
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 24

    # 保存
    wb.save(output_path)
    print(f"台账已生成: {output_path}")
    print(f"  期数: {len(data.get('periods', []))}")
    print(f"  合计利息: {format_cny(total_interest)}")


def main():
    parser = argparse.ArgumentParser(description="生成工程款利息核算台账")
    parser.add_argument("--input", "-i", required=True, help="输入 JSON 数据文件")
    parser.add_argument("--output", "-o", required=True, help="输出 .xlsx 文件路径")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    create_ledger(data, args.output)


if __name__ == "__main__":
    main()
